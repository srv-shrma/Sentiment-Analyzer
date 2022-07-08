# USED LIBRARIES
import requests
from bs4 import BeautifulSoup
import os
import pandas as pd
from textblob import TextBlob
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import re
import textstat
from openpyxl import load_workbook


# FUNCTIONS
def per_comp_words(s):
    words = s.split()
    return ((comp_words(s)/len(words)) * 100)

def comp_words(s):
    english_most_common_10k = 'https://raw.githubusercontent.com/first20hours/google-10000-english/master/google-10000-english-usa-no-swears.txt'
    # Get the file of 10 k most common words from TXT file in a github repo
    response = requests.get(english_most_common_10k)
    data = response.text
    set_of_common_words = {x for x in data.split('\n')}
    ses = s.split(' ')
    cnt = 0
    for word in ses:
        # word = input()
        if word in set_of_common_words:
            # print(f'The word "{word}" is common')
            pass
        else:
            cnt += 1
            # print(f'The word "{word}" is difficult')
    return cnt

def avg_no_words_sen(s):
    parts = [len(l.split()) for l in re.split(r'[?!.]', s) if l.strip()]
    return (sum(parts) / len(parts))

def avg_sen_len(s):
    sents = s.split('.')
    return sum(len(x.split()) for x in sents) / len(sents)

analyser = SentimentIntensityAnalyzer()
def print_sentiment_scores(tweets):
    vadersenti = analyser.polarity_scores(tweets)
    # return pd.Series([vadersenti['pos'], vadersenti['neg'], vadersenti['neu'], vadersenti['compound']])
    a = list([vadersenti['pos'], vadersenti['neg'], vadersenti['neu'], vadersenti['compound']])
    return a

def syllable_count(word):
    word = word.lower()
    count = 0
    vowels = "aeiouy"
    if word[0] in vowels:
        count += 1
    for index in range(1, len(word)):
        if word[index] in vowels and word[index - 1] not in vowels:
            count += 1
    if word.endswith("e"):
        count -= 1
    if count == 0:
        count += 1
    return count

def per_pronouns(s):
    words = s.split()
    lst = ['i', 'he', 'him', 'her', 'it', 'me', 'she', 'them', 'they', 'us', 'we', 'you']
    cnt = 0
    for word in words:
        if word.lower() in lst:
            cnt += 1
    return cnt

def avg_word_len(s):
    words = s.split()
    average = sum(len(word) for word in words) / len(words)
    return average

def get_content(doc):
    selection_class = 'td-post-content'
    title = doc.find_all('h1', {'class': 'entry-title'})
    article = doc.find_all('div', {'class': selection_class})
    # print(title[0].text.strip())
    tlst = title[0].text.strip().split()
    t_path = ""
    t_path = t_path.join(tlst)
    lst = [title[0].text.strip(), article[0].text.strip()]
    st = ' '
    st = st.join(lst)
    s = "\n"
    s = s.join(lst)
    p = "data/{}.txt".format(t_path[:-1])
    if os.path.exists(p):
        print('The file {} already exists. Skipping...'.format(p))
    elif t_path.find('/') != -1 or t_path.find('(') != -1:
        print('The file {} '.format(p))
        pass
    else:
        # file = open(p, "w")
        # file.write(s)
        with open(p, "w", encoding="utf-8") as file:
            file.write(s)
        # file.flush()
        file.close()

    return st

def get_urls_page(df):
    url_id = 1
    # url = df['URL'][0]
    # response = requests.get(url, headers={"User-Agent": "XY"})
    # if response.status_code != 200:
    #     raise Exception('Failed to load page {}'.format(url))
    # doc = BeautifulSoup(response.text, 'html.parser')
    # text_dict[url_id] = get_content(doc)
    for url in df['URL']:
        response = requests.get(url, headers={"User-Agent": "XY"})
        if response.status_code != 200:
            raise Exception('Failed to load page {}'.format(url))
        doc = BeautifulSoup(response.text, 'html.parser')
        text_dict[url_id] = get_content(doc)
        url_id += 1

# DRIVER CODE
if __name__ == '__main__':
    df = pd.read_excel(r'C:\Users\offic\PycharmProjects\pythonProject\Input.xlsx')
    # print(df['URL'][0])
    text_dict = {}
    get_urls_page(df)
    print(len(text_dict))

    wb = load_workbook(filename='Output.xlsx')
    s = wb.active
    # print(s.cell(2, 2).value)
    for i in range(0, len(text_dict)):
        temp = text_dict[i+1]
        words = temp.split()
        sen = temp.split('.')
        avg = sum(syllable_count(word) for word in words) / len(words)
        pos_neg = print_sentiment_scores(temp)
        s.cell(2+i, 3).value = pos_neg[0]
        s.cell(2+i, 4).value = pos_neg[1]
        s.cell(2+i, 5).value = TextBlob(temp).sentiment.polarity
        s.cell(2+i, 6).value = TextBlob(temp).sentiment.subjectivity
        s.cell(2+i, 7).value = avg_sen_len(temp)
        s.cell(2+i, 8).value = per_comp_words(temp)
        s.cell(2+i, 9).value = textstat.gunning_fog(temp)
        s.cell(2+i, 10).value = avg_no_words_sen(temp)
        s.cell(2+i, 11).value = comp_words(temp)
        s.cell(2+i, 12).value = len(words)
        s.cell(2+i, 13).value = avg
        s.cell(2+i, 14).value = per_pronouns(temp)
        s.cell(2+i, 15).value = avg_word_len(temp)

    wb.save("Output.xlsx")

    # print('personal pronouns - {}'.format(per_pronouns(text_dict[1])))
    # print('avg word len - {}'.format(avg_word_len(text_dict[1])))
    # words = text_dict[1].split()
    # sen = text_dict[1].split('.')
    # avg = sum(syllable_count(word) for word in words) / len(words)
    # print('avg syllable per word - {}'.format(avg))
    # print('word count - {}'.format(len(words)))
    # print(TextBlob(text_dict[1]).sentiment)
    # print(TextBlob(text_dict[1]).sentiment.polarity)
    # print(TextBlob(text_dict[1]).sentiment.subjectivity)
    # print(print_sentiment_scores(text_dict[1]))
    # print('avg sen len - {}'.format(avg_sen_len(text_dict[1])))
    # print('avg no of words per sen - {}'.format(avg_no_words_sen(text_dict[1])))
    # print('fog index - {}'.format(textstat.gunning_fog(text_dict[1])))
    #
    # # 0.4[(words / sentences) + 100(complex words / words)].
    # print(0.4 * ((len(words)/len(sen)) + 100 * (comp_words(text_dict[1])/len(words))))
    #
    # print('complex words - {}'.format(comp_words(text_dict[1])))
    # print('per complex words - {}'.format(per_comp_words(text_dict[1])))



